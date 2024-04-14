import calendar
from datetime import date

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook

polnum = '07329178'
str_date = date(2031, 1, 30)
end_date = date(2032, 6, 30)
curr_date = str_date
tmp_day1 = str_date.day

# new excel file
wb = Workbook()
ws = wb.active
# ws.delete_cols(1, 3)  # delete 3 columns

i = 1
j = 1
while curr_date <= end_date:
    print(curr_date)
    ws.cell(row=i, column=1).value = 'l2polrnwl'
    ws.cell(row=i, column=2).value = curr_date
    ws.cell(row=i, column=3).value = polnum
    i += 1
    ws.cell(row=i, column=1).value = 'l2newunitd'
    ws.cell(row=i, column=2).value = curr_date
    ws.cell(row=i, column=3).value = polnum
    i += 1
    curr_date = curr_date + relativedelta(months=1)

    try:
        curr_date = curr_date.replace(day=tmp_day1)
    except ValueError:
        last_day = calendar.monthrange(curr_date.year, curr_date.month)[1]
        curr_date = curr_date.replace(day=last_day)

xl_filename = "jobs" + '-' + polnum + '-' + str(str_date) + '-' + str(end_date) + ".xlsx"
wb.save(xl_filename)
