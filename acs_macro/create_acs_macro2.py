from pathlib import Path

import openpyxl


def get_screen_xml(screen_num, entry_screen_flag, exit_screen_flag, input_val):
    ha_script = f"""
    <screen name="Screen{screen_num}" entryscreen="{entry_screen_flag}" exitscreen="{exit_screen_flag}" transient="false">
        <description>
            <oia status="NOTINHIBITED" optional="false" invertmatch="false"/>
            <numfields number="101" optional="true" invertmatch="false"/>
            <numinputfields number="4" optional="true" invertmatch="false"/>
        </description>

        <actions>
            <input value="{input_val}" row="0" col="0" movecursor="true" xlatehostkeys="true"
                   encrypted="false"/>
        </actions>

        <nextscreens timeout="0">
            <nextscreen name="Screen{screen_num + 1}"/>
        </nextscreens>
    </screen>
    """

    return ha_script


def write_it(outf, job, eff_date, pol, row, last_row):
    screen_num = (row - 1) * 3
    tab_key = "" if len(job) == 10 else "[tab]"

    entry_screen_flag = "true" if row == 1 else "false"
    exit_screen_flag = "false"
    input_val = f'{job}{tab_key}{eff_date}[enter]'
    ha_script = get_screen_xml(screen_num + 1, entry_screen_flag, exit_screen_flag, input_val)
    outf.write(ha_script)

    entry_screen_flag = "false"
    exit_screen_flag = "false"
    input_val = f'{pol}{pol}[enter]'
    ha_script = get_screen_xml(screen_num + 2, entry_screen_flag, exit_screen_flag, input_val)
    outf.write(ha_script)

    entry_screen_flag = "false"
    exit_screen_flag = "true" if last_row else "false"
    input_val = '[tab]y[enter]'
    ha_script = get_screen_xml(screen_num + 3, entry_screen_flag, exit_screen_flag, input_val)
    outf.write(ha_script)


def gen_acs_macro():
    xl_filename = Path('jobs.xlsx')

    wb1 = openpyxl.load_workbook(xl_filename)
    wb = wb1.active

    # new macro file
    macro_filename = xl_filename.stem.replace('jobs', 'macro') + '.mac'

    with open(macro_filename, 'w') as outf:
        # ha script header
        outf.write(
            '<HAScript name="macro1" description="" timeout="60000" pausetime="300" promptall="true" blockinput="true" author="vpc" creationdate="Nov 15, 2023 8:21:30 AM" supressclearevents="false" usevars="false" ignorepauseforenhancedtn="true" delayifnotenhancedtn="0" ignorepausetimeforenhancedtn="true" continueontimeout="false">')

        # ha script details (screen sequence)
        for row in range(1, wb.max_row + 1):
            job = wb.cell(row, 1).value
            eff_date_raw = str(wb.cell(row, 2).value)
            eff_date = f"{eff_date_raw[5:7]}/{eff_date_raw[8:10]}/{eff_date_raw[0:4]}"  # convert YYYY-MM-DD to MM/DD/YYYY format
            pol = wb.cell(row, 3).value
            write_it(outf, job, eff_date, pol, row, row == wb.max_row)

        # ha script footer
        outf.write('</HAScript>')


if __name__ == '__main__':
    gen_acs_macro()
