from datetime import date

import openpyxl
from dateutil.relativedelta import relativedelta

polnum = '07329178'
str_date = date(2031, 6, 17)
end_date = date(2032, 6, 17)
curr_date = str_date

# xl_filename = "jobs" + polnum + str(str_date) + str(end_date) + ".xlsx"
xl_filename = "jobs.xlsx"
wb1 = openpyxl.load_workbook(xl_filename)
ws1 = wb1.worksheets[0]
# ws1.delete_cols(1,3) # delete 3 columns

i = 1
j = 1
while curr_date <= end_date:
    print(curr_date)
    ws1.cell(row=i, column=1).value = 'l2polrnwl'
    ws1.cell(row=i, column=2).value = curr_date
    ws1.cell(row=i, column=3).value = polnum
    i += 1
    ws1.cell(row=i, column=1).value = 'l2newunitd'
    ws1.cell(row=i, column=2).value = curr_date
    ws1.cell(row=i, column=3).value = polnum
    i += 1
    curr_date = curr_date + relativedelta(months=1)

wb1.save(xl_filename)
