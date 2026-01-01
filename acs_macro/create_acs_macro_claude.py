"""
ACS Macro Generator
Generates Host Access Scripts from Excel job data for automated terminal emulator interactions.
"""

import openpyxl
from typing import TextIO
from datetime import datetime
from pathlib import Path

# Constants
COL_JOB = 1
COL_EFF_DATE = 2
COL_POLICY = 3

SCREENS_PER_ROW = 3
JOB_LENGTH_THRESHOLD = 10

DEFAULT_TIMEOUT = "0"
DEFAULT_PAUSE_TIME = "300"
SCRIPT_TIMEOUT = "60000"


class MacroGenerationError(Exception):
    """Custom exception for macro generation errors."""
    pass


def format_effective_date(date_value) -> str:
    """
    Convert date value to MM/DD/YYYY format.

    Args:
        date_value: Date value from Excel (datetime or string)

    Returns:
        Formatted date string in MM/DD/YYYY format

    Raises:
        ValueError: If date format is invalid
    """
    try:
        if isinstance(date_value, datetime):
            return date_value.strftime("%m/%d/%Y")

        date_str = str(date_value)
        # Assuming format: YYYY-MM-DD
        if len(date_str) >= 10:
            return f"{date_str[5:7]}/{date_str[8:10]}/{date_str[0:4]}"

        raise ValueError(f"Invalid date format: {date_value}")
    except Exception as e:
        raise ValueError(f"Error formatting date '{date_value}': {str(e)}")


def get_screen_xml(screen_num: int, entry_screen: bool, exit_screen: bool,
                   input_value: str, num_fields: int, num_input_fields: int,
                   next_screen_num: int = None) -> str:
    """
    Generate XML for a single screen definition.

    Args:
        screen_num: Screen number identifier
        entry_screen: Whether this is an entry screen
        exit_screen: Whether this is an exit screen
        input_value: Input value to send to the screen
        num_fields: Number of fields on the screen
        num_input_fields: Number of input fields
        next_screen_num: Next screen number (None if exit screen)

    Returns:
        XML string for the screen
    """
    entry_flag = "true" if entry_screen else "false"
    exit_flag = "true" if exit_screen else "false"

    next_screen_xml = ""
    if next_screen_num is not None:
        next_screen_xml = f"""
        <nextscreens timeout="{DEFAULT_TIMEOUT}">
            <nextscreen name="Screen{next_screen_num}"/>
        </nextscreens>"""

    return f"""
    <screen name="Screen{screen_num}" entryscreen="{entry_flag}" exitscreen="{exit_flag}" transient="false">
        <description>
            <oia status="NOTINHIBITED" optional="false" invertmatch="false"/>
            <numfields number="{num_fields}" optional="true" invertmatch="false"/>
            <numinputfields number="{num_input_fields}" optional="true" invertmatch="false"/>
        </description>
        <actions>
            <input value="{input_value}" row="0" col="0" movecursor="true" xlatehostkeys="true"
                   encrypted="false"/>
        </actions>{next_screen_xml}
    </screen>"""


def write_screen_sequence(outf: TextIO, job: str, eff_date: str, policy: str,
                          row_num: int, is_last_row: bool) -> None:
    """
    Write the three-screen sequence for a single job entry.

    Args:
        outf: Output file handle
        job: Job identifier
        eff_date: Effective date in MM/DD/YYYY format
        policy: Policy number
        row_num: Current row number (1-indexed)
        is_last_row: Whether this is the last row
    """
    is_first_row = (row_num == 1)
    base_screen_num = (row_num - 1) * SCREENS_PER_ROW

    # Determine if tab key is needed based on job length
    tab_key = "" if len(job) == JOB_LENGTH_THRESHOLD else "[tab]"

    # Screen 1: Job and effective date entry
    screen1_xml = get_screen_xml(
        screen_num=base_screen_num + 1,
        entry_screen=is_first_row,
        exit_screen=False,
        input_value=f"{job}{tab_key}{eff_date}[enter]",
        num_fields=101,
        num_input_fields=4,
        next_screen_num=base_screen_num + 2
    )

    # Screen 2: Policy entry
    screen2_xml = get_screen_xml(
        screen_num=base_screen_num + 2,
        entry_screen=False,
        exit_screen=False,
        input_value=f"{policy}{policy}[enter]",
        num_fields=36,
        num_input_fields=2,
        next_screen_num=base_screen_num + 3
    )

    # Screen 3: Confirmation
    screen3_xml = get_screen_xml(
        screen_num=base_screen_num + 3,
        entry_screen=False,
        exit_screen=is_last_row,
        input_value="[tab]y[enter]",
        num_fields=36,
        num_input_fields=1,
        next_screen_num=None if is_last_row else base_screen_num + 4
    )

    outf.write(screen1_xml)
    outf.write(screen2_xml)
    outf.write(screen3_xml)


def get_macro_header() -> str:
    """Generate the HAScript header XML."""
    return (
        '<HAScript name="macro1" description="" '
        f'timeout="{SCRIPT_TIMEOUT}" pausetime="{DEFAULT_PAUSE_TIME}" '
        'promptall="true" blockinput="true" author="vpc" '
        'creationdate="Nov 15, 2023 8:21:30 AM" supressclearevents="false" '
        'usevars="false" ignorepauseforenhancedtn="true" '
        'delayifnotenhancedtn="0" ignorepausetimeforenhancedtn="true" '
        'continueontimeout="false">'
    )


def get_macro_footer() -> str:
    """Generate the HAScript footer XML."""
    return '</HAScript>'


def validate_row_data(job, eff_date, policy, row_num: int) -> None:
    """
    Validate data from a spreadsheet row.

    Args:
        job: Job identifier
        eff_date: Effective date
        policy: Policy number
        row_num: Row number for error reporting

    Raises:
        MacroGenerationError: If data is invalid
    """
    if not job:
        raise MacroGenerationError(f"Row {row_num}: Job identifier is missing")

    if not eff_date:
        raise MacroGenerationError(f"Row {row_num}: Effective date is missing")

    if not policy:
        raise MacroGenerationError(f"Row {row_num}: Policy number is missing")


def generate_acs_macro(input_filename: str = 'jobs.xlsx') -> str:
    """
    Generate ACS macro file from Excel job data.

    Args:
        input_filename: Path to input Excel file

    Returns:
        Path to generated macro file

    Raises:
        MacroGenerationError: If generation fails
        FileNotFoundError: If input file doesn't exist
    """
    # Validate input file exists
    input_path = Path(input_filename)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_filename}")

    # Generate output filename
    output_filename = input_path.stem.replace('jobs', 'macro') + '.mac'

    try:
        # Load workbook
        workbook = openpyxl.load_workbook(input_filename)
        worksheet = workbook.active

        if worksheet.max_row < 1:
            raise MacroGenerationError("Workbook is empty")

        # Write macro file
        with open(output_filename, 'w', encoding='utf-8') as outf:
            # Write header
            outf.write(get_macro_header())

            # Process each row
            for row_idx in range(1, worksheet.max_row + 1):
                # Extract data from columns
                job = worksheet.cell(row_idx, COL_JOB).value
                eff_date_raw = worksheet.cell(row_idx, COL_EFF_DATE).value
                policy = worksheet.cell(row_idx, COL_POLICY).value

                # Validate row data
                validate_row_data(job, eff_date_raw, policy, row_idx)

                # Format effective date
                eff_date = format_effective_date(eff_date_raw)

                # Write screen sequence for this row
                is_last_row = (row_idx == worksheet.max_row)
                write_screen_sequence(outf, job, eff_date, policy, row_idx, is_last_row)

            # Write footer
            outf.write(get_macro_footer())

        print(f"Successfully generated macro file: {output_filename}")
        return output_filename

    except openpyxl.utils.exceptions.InvalidFileException as e:
        raise MacroGenerationError(f"Invalid Excel file format: {str(e)}")
    except Exception as e:
        raise MacroGenerationError(f"Error generating macro: {str(e)}")


def main():
    """Main entry point for the script."""
    try:
        output_file = generate_acs_macro()
        print(f"Macro generation complete: {output_file}")
    except FileNotFoundError as e:
        print(f"Error: {e}")
        return 1
    except MacroGenerationError as e:
        print(f"Macro generation failed: {e}")
        return 1
    except Exception as e:
        print(f"Unexpected error: {e}")
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
